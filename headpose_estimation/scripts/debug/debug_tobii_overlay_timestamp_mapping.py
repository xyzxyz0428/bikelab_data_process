#!/usr/bin/env python3
import argparse
import csv
import json
from pathlib import Path
from datetime import datetime, timezone

import cv2
import numpy as np
from openpyxl import load_workbook


def iso_to_unix_ns(iso_str: str) -> int:
    s = str(iso_str).strip()
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    dt = datetime.fromisoformat(s)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return int(round(dt.timestamp() * 1e9))


def get_recording_start_ns(recording_g3_path):
    with open(recording_g3_path, "r", encoding="utf-8") as f:
        g3 = json.load(f)
    return iso_to_unix_ns(g3["created"])


def read_csv_dicts(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def to_float(v, default=np.nan):
    try:
        if v is None or v == "":
            return default
        return float(v)
    except Exception:
        return default


def to_int(v, default=None):
    try:
        if v is None or v == "":
            return default
        return int(float(v))
    except Exception:
        return default


def resolve_frame_path(frame_dir, frame_idx):
    for ext in [".png", ".jpg", ".jpeg"]:
        p = frame_dir / f"frame_{int(frame_idx):06d}{ext}"
        if p.exists():
            return p
    return None


def load_scene_timestamps(path):
    rows = read_csv_dicts(path)
    out = []

    for r in rows:
        frame_idx = to_int(r.get("frame_idx"))
        unix_ns = to_int(r.get("unix_ns"))

        if frame_idx is None or unix_ns is None:
            continue

        out.append({
            "frame_idx": frame_idx,
            "unix_ns": unix_ns,
        })

    return out


def load_tobii_xlsx(xlsx_path, recording_g3_path):
    recording_start_ns = get_recording_start_ns(recording_g3_path)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = ["" if v is None else str(v).strip() for v in next(rows_iter)]
    hidx = {k: i for i, k in enumerate(header)}

    required = [
        "Recording timestamp",
        "Computer timestamp",
        "Sensor",
        "Fixation point X",
        "Fixation point Y",
    ]

    for k in required:
        if k not in hidx:
            raise KeyError(f"Missing column in xlsx: {k}")

    media_w_col = hidx.get("Recording media width", None)
    media_h_col = hidx.get("Recording media height", None)

    out = []
    first_media_size = None

    for row in rows_iter:
        sensor = row[hidx["Sensor"]]

        # 一般只用 Eye Tracker 行
        if sensor != "Eye Tracker":
            continue

        gx = to_float(row[hidx["Fixation point X"]])
        gy = to_float(row[hidx["Fixation point Y"]])

        if not np.isfinite(gx) or not np.isfinite(gy):
            continue

        rec_us = to_float(row[hidx["Recording timestamp"]])
        comp_us = to_float(row[hidx["Computer timestamp"]])

        if not np.isfinite(rec_us) or not np.isfinite(comp_us):
            continue

        rec_unix_ns = recording_start_ns + int(round(rec_us * 1000.0))
        comp_unix_ns = recording_start_ns + int(round(comp_us * 1000.0))

        if media_w_col is not None and media_h_col is not None:
            mw = row[media_w_col]
            mh = row[media_h_col]
            if first_media_size is None and mw not in [None, ""] and mh not in [None, ""]:
                first_media_size = (mw, mh)

        out.append({
            "rec_unix_ns": rec_unix_ns,
            "comp_unix_ns": comp_unix_ns,
            "recording_timestamp_us": rec_us,
            "computer_timestamp_us": comp_us,
            "x": gx,
            "y": gy,
        })

    out_rec_sorted = sorted(out, key=lambda r: r["rec_unix_ns"])
    out_comp_sorted = sorted(out, key=lambda r: r["comp_unix_ns"])

    return out_rec_sorted, out_comp_sorted, first_media_size, recording_start_ns


def nearest_sample(samples, target_ns, time_key):
    if not samples:
        return None, None

    # simple binary search
    times = [s[time_key] for s in samples]
    idx = np.searchsorted(times, target_ns)

    candidates = []
    if 0 <= idx < len(samples):
        candidates.append(idx)
    if 0 <= idx - 1 < len(samples):
        candidates.append(idx - 1)

    if not candidates:
        return None, None

    best_i = min(candidates, key=lambda i: abs(samples[i][time_key] - target_ns))
    return samples[best_i], abs(samples[best_i][time_key] - target_ns) / 1e6


def interpolate_sample(samples, target_ns, time_key):
    if len(samples) < 2:
        return None, None

    times = [s[time_key] for s in samples]
    idx = np.searchsorted(times, target_ns)

    if idx <= 0 or idx >= len(samples):
        return None, None

    s0 = samples[idx - 1]
    s1 = samples[idx]

    t0 = s0[time_key]
    t1 = s1[time_key]

    if t1 == t0:
        return None, None

    a = (target_ns - t0) / (t1 - t0)

    x = (1 - a) * s0["x"] + a * s1["x"]
    y = (1 - a) * s0["y"] + a * s1["y"]

    dt_ms = min(abs(target_ns - t0), abs(target_ns - t1)) / 1e6

    return {
        "x": x,
        "y": y,
        "t0_ns": t0,
        "t1_ns": t1,
        "alpha": a,
    }, dt_ms


def draw_point(img, x, y, color, label):
    if not np.isfinite(x) or not np.isfinite(y):
        return

    x = int(round(x))
    y = int(round(y))

    cv2.circle(img, (x, y), 10, color, 2, cv2.LINE_AA)
    cv2.drawMarker(
        img,
        (x, y),
        color,
        markerType=cv2.MARKER_CROSS,
        markerSize=28,
        thickness=2,
        line_type=cv2.LINE_AA,
    )
    cv2.putText(
        img,
        label,
        (x + 12, y - 12),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.55,
        color,
        2,
        cv2.LINE_AA,
    )


def load_validation_centers(path):
    if path is None:
        return {}

    rows = read_csv_dicts(path)
    centers_by_frame = {}

    for r in rows:
        frame_idx = None
        for k in ["frame_idx", "scene_frame_idx", "nearest_scene_frame_idx"]:
            if k in r:
                frame_idx = to_int(r.get(k))
                break

        if frame_idx is None:
            continue

        cx = np.nan
        cy = np.nan

        for k in ["tag_center_x", "center_x", "tag_center_u", "tag_u"]:
            if k in r:
                cx = to_float(r.get(k))
                break

        for k in ["tag_center_y", "center_y", "tag_center_v", "tag_v"]:
            if k in r:
                cy = to_float(r.get(k))
                break

        if np.isfinite(cx) and np.isfinite(cy):
            centers_by_frame.setdefault(frame_idx, []).append((cx, cy, r.get("tag_id", "")))

    return centers_by_frame


def choose_frames(scene_rows, frame_indices, every_n, max_frames):
    if frame_indices:
        wanted = set(frame_indices)
        return [r for r in scene_rows if r["frame_idx"] in wanted]

    selected = scene_rows[::every_n]

    if max_frames is not None and len(selected) > max_frames:
        selected = selected[:max_frames]

    return selected


def main():
    ap = argparse.ArgumentParser()

    ap.add_argument("--tobii-xlsx", required=True)
    ap.add_argument("--recording-g3", required=True)
    ap.add_argument("--scene-timestamps-csv", required=True)
    ap.add_argument("--scene-frame-dir", required=True)
    ap.add_argument("--output-dir", required=True)

    ap.add_argument("--validation-csv", default=None,
                    help="optional tobii_2d_validation csv to draw tag centers")

    ap.add_argument("--frame-indices", default="",
                    help="comma-separated frame indices to visualize")
    ap.add_argument("--every-n", type=int, default=100)
    ap.add_argument("--max-frames", type=int, default=30)

    args = ap.parse_args()

    scene_rows = load_scene_timestamps(args.scene_timestamps_csv)
    scene_frame_dir = Path(args.scene_frame_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    tobii_rec, tobii_comp, media_size, recording_start_ns = load_tobii_xlsx(
        args.tobii_xlsx,
        args.recording_g3,
    )

    centers_by_frame = load_validation_centers(args.validation_csv)

    frame_indices = []
    if args.frame_indices.strip():
        frame_indices = [int(x) for x in args.frame_indices.split(",") if x.strip()]

    selected_scene_rows = choose_frames(
        scene_rows=scene_rows,
        frame_indices=frame_indices,
        every_n=args.every_n,
        max_frames=args.max_frames,
    )

    report_rows = []

    print("[INFO] recording_start_ns:", recording_start_ns)
    print("[INFO] Tobii rows by Recording timestamp:", len(tobii_rec))
    print("[INFO] Tobii rows by Computer timestamp:", len(tobii_comp))
    print("[INFO] first media size from xlsx:", media_size)
    print("[INFO] selected scene frames:", len(selected_scene_rows))

    for sr in selected_scene_rows:
        frame_idx = sr["frame_idx"]
        scene_ns = sr["unix_ns"]

        frame_path = resolve_frame_path(scene_frame_dir, frame_idx)
        if frame_path is None:
            continue

        img = cv2.imread(str(frame_path), cv2.IMREAD_COLOR)
        if img is None:
            continue

        comp_near, comp_dt = nearest_sample(tobii_comp, scene_ns, "comp_unix_ns")
        rec_near, rec_dt = nearest_sample(tobii_rec, scene_ns, "rec_unix_ns")
        rec_interp, rec_interp_dt = interpolate_sample(tobii_rec, scene_ns, "rec_unix_ns")

        # draw tag centers if available
        for cx, cy, tag_id in centers_by_frame.get(frame_idx, []):
            draw_point(img, cx, cy, (0, 255, 255), f"tag {tag_id}")

        if comp_near is not None:
            draw_point(img, comp_near["x"], comp_near["y"], (0, 0, 255), "comp nearest")

        if rec_near is not None:
            draw_point(img, rec_near["x"], rec_near["y"], (255, 0, 0), "rec nearest")

        if rec_interp is not None:
            draw_point(img, rec_interp["x"], rec_interp["y"], (0, 255, 0), "rec interp")

        header1 = f"frame={frame_idx} scene_ns={scene_ns}"
        header2 = (
            f"comp_dt={comp_dt:.2f}ms | rec_dt={rec_dt:.2f}ms | "
            f"interp_nearest_dt={rec_interp_dt:.2f}ms"
            if comp_dt is not None and rec_dt is not None and rec_interp_dt is not None
            else ""
        )

        cv2.rectangle(img, (0, 0), (img.shape[1], 60), (0, 0, 0), -1)
        cv2.putText(img, header1, (10, 22), cv2.FONT_HERSHEY_SIMPLEX, 0.6,
                    (255, 255, 255), 2, cv2.LINE_AA)
        cv2.putText(img, header2, (10, 48), cv2.FONT_HERSHEY_SIMPLEX, 0.6,
                    (255, 255, 255), 2, cv2.LINE_AA)

        out_path = output_dir / f"overlay_debug_frame_{frame_idx:06d}.png"
        cv2.imwrite(str(out_path), img)

        report_rows.append({
            "frame_idx": frame_idx,
            "scene_unix_ns": scene_ns,

            "comp_nearest_x": comp_near["x"] if comp_near else "",
            "comp_nearest_y": comp_near["y"] if comp_near else "",
            "comp_nearest_dt_ms": comp_dt if comp_dt is not None else "",

            "rec_nearest_x": rec_near["x"] if rec_near else "",
            "rec_nearest_y": rec_near["y"] if rec_near else "",
            "rec_nearest_dt_ms": rec_dt if rec_dt is not None else "",

            "rec_interp_x": rec_interp["x"] if rec_interp else "",
            "rec_interp_y": rec_interp["y"] if rec_interp else "",
            "rec_interp_dt_ms": rec_interp_dt if rec_interp_dt is not None else "",

            "output_image": str(out_path),
        })

    report_csv = output_dir / "overlay_debug_report.csv"
    with open(report_csv, "w", encoding="utf-8", newline="") as f:
        if report_rows:
            writer = csv.DictWriter(f, fieldnames=list(report_rows[0].keys()))
            writer.writeheader()
            writer.writerows(report_rows)

    print("[INFO] saved images to:", output_dir)
    print("[INFO] saved report to:", report_csv)
    print()
    print("Color legend:")
    print("  yellow = tag center, if validation csv is provided")
    print("  red    = Computer timestamp + nearest")
    print("  blue   = Recording timestamp + nearest")
    print("  green  = Recording timestamp + interpolation")


if __name__ == "__main__":
    main()