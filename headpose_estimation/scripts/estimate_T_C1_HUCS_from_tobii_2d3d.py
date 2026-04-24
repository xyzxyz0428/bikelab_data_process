#!/usr/bin/env python3
import argparse
import json
from datetime import datetime, timezone

import cv2
import numpy as np
from openpyxl import load_workbook


def iso_to_unix_ns(iso_str: str) -> int:
    s = iso_str.strip()
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    dt = datetime.fromisoformat(s)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return int(round(dt.timestamp() * 1e9))


def get_recording_start_ns(recording_g3_path):
    with open(recording_g3_path, "r", encoding="utf-8") as f:
        g3 = json.load(f)
    return iso_to_unix_ns(g3["created"])


def load_camera_json(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    K = np.array(data["K"], dtype=np.float64)
    dist = np.array(data["dist"], dtype=np.float64).reshape(-1, 1)
    return K, dist


def rt_to_T(Rm, t):
    T = np.eye(4, dtype=np.float64)
    T[:3, :3] = Rm
    T[:3, 3] = np.asarray(t).reshape(3)
    return T


def to_float(v, default=np.nan):
    try:
        if v is None or v == "":
            return default
        return float(v)
    except Exception:
        return default


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tobii-xlsx", required=True)
    ap.add_argument("--recording-g3", required=True)
    ap.add_argument("--scene-camera-json", required=True)
    ap.add_argument("--output-json", required=True)
    ap.add_argument("--max-samples", type=int, default=5000)
    args = ap.parse_args()

    K, dist = load_camera_json(args.scene_camera_json)
    _ = get_recording_start_ns(args.recording_g3)  # not strictly needed here, but kept for consistency

    wb = load_workbook(args.tobii_xlsx, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = ["" if v is None else str(v).strip() for v in rows[0]]
    hidx = {k: i for i, k in enumerate(header)}

    required = [
        "Sensor",
        "Gaze point X", "Gaze point Y",
        "Gaze point 3D X", "Gaze point 3D Y", "Gaze point 3D Z"
    ]
    for k in required:
        if k not in hidx:
            raise RuntimeError(f"Missing column: {k}")

    obj_pts = []
    img_pts = []

    count_total_eye = 0
    count_valid_2d3d = 0
    count_missing_2d = 0
    count_missing_3d = 0

    for row in rows[1:]:
        sensor = row[hidx["Sensor"]]
        if sensor != "Eye Tracker":
            continue

        count_total_eye += 1

        gx = to_float(row[hidx["Gaze point X"]])
        gy = to_float(row[hidx["Gaze point Y"]])
        x3 = to_float(row[hidx["Gaze point 3D X"]])
        y3 = to_float(row[hidx["Gaze point 3D Y"]])
        z3 = to_float(row[hidx["Gaze point 3D Z"]])

        if any(np.isnan(v) for v in [gx, gy]):
            count_missing_2d += 1
            continue

        if any(np.isnan(v) for v in [x3, y3, z3]):
            count_missing_3d += 1
            continue

        obj_pts.append([x3 * 1e-3, y3 * 1e-3, z3 * 1e-3])  # HUCS mm -> m
        img_pts.append([gx, gy])

        count_valid_2d3d += 1
        if count_valid_2d3d >= args.max_samples:
            break

    print(f"total Eye Tracker rows = {count_total_eye}")
    print(f"valid 2D+3D rows = {count_valid_2d3d}")
    print(f"missing 2D rows = {count_missing_2d}")
    print(f"missing 3D rows = {count_missing_3d}")

    obj_pts = np.asarray(obj_pts, dtype=np.float64)
    img_pts = np.asarray(img_pts, dtype=np.float64)

    if len(obj_pts) < 20:
        raise RuntimeError("Too few valid Tobii 2D/3D samples.")

    ok, rvec, tvec, inliers = cv2.solvePnPRansac(
        objectPoints=obj_pts,
        imagePoints=img_pts,
        cameraMatrix=K,
        distCoeffs=dist,
        iterationsCount=500,
        reprojectionError=8.0,
        confidence=0.999,
        flags=cv2.SOLVEPNP_EPNP,
    )
    if not ok:
        ok, rvec, tvec = cv2.solvePnP(
            obj_pts, img_pts, K, dist, flags=cv2.SOLVEPNP_ITERATIVE
        )
        if not ok:
            raise RuntimeError("solvePnP failed")

    try:
        if inliers is not None and len(inliers) >= 10:
            idx = inliers.reshape(-1)
            obj_ref = obj_pts[idx]
            img_ref = img_pts[idx]
        else:
            obj_ref = obj_pts
            img_ref = img_pts
        rvec, tvec = cv2.solvePnPRefineLM(obj_ref, img_ref, K, dist, rvec, tvec)
    except cv2.error:
        pass

    Rm, _ = cv2.Rodrigues(rvec)
    T_C1_HUCS = rt_to_T(Rm, tvec.reshape(3))

    result = {
        "num_samples": int(len(obj_pts)),
        "num_inliers": int(len(inliers)) if inliers is not None else 0,
        "T_C1_HUCS": T_C1_HUCS.tolist()
    }

    with open(args.output_json, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2)

    print(f"saved to {args.output_json}")


if __name__ == "__main__":
    main()