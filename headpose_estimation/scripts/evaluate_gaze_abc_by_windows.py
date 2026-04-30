#!/usr/bin/env python3
import argparse
import csv
import json
from datetime import datetime, timezone
from collections import defaultdict
from pathlib import Path
import os
import sys
import gc

import numpy as np
from openpyxl import load_workbook
from scipy.spatial.transform import Rotation as R


def iso_to_unix_ns(iso_str: str) -> int:
    s = iso_str.strip()
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    dt = datetime.fromisoformat(s)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return int(round(dt.timestamp() * 1e9))


def get_recording_start_ns(recording_g3_path):
    with open(recording_g3_path, "r", encoding="utf-8") as f:
        g3 = json.load(f)
    return iso_to_unix_ns(g3["created"])


def read_csv_dicts(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def to_int(v, default=None):
    try:
        if v is None or v == "":
            return default
        return int(float(v))
    except Exception:
        return default


def to_float(v, default=np.nan):
    try:
        if v is None or v == "":
            return default
        return float(v)
    except Exception:
        return default


def load_camera_json(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    K = np.array(data["K"], dtype=np.float64)
    dist = np.array(data["dist"], dtype=np.float64).reshape(-1, 1)
    return K, dist


def load_baseline_json(
    path,
    exclude_low_confidence=True,
    max_tag_translation_std_m=None,
    max_tag_rotation_std_deg=None,
):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    tags = {}
    debug_info = {
        "all_tag_ids": [],
        "kept_tag_ids": [],
        "removed": {},
    }

    for sid, item in data["tags"].items():
        tid = int(sid)
        debug_info["all_tag_ids"].append(tid)

        reasons = []

        low_conf = bool(item.get("low_confidence", False))
        trans_std = float(item.get("translation_std_m", 0.0))
        rot_std = float(item.get("rotation_std_deg", 0.0))

        if exclude_low_confidence and low_conf:
            reasons.append("low_confidence")

        if max_tag_translation_std_m is not None:
            if trans_std > max_tag_translation_std_m:
                reasons.append(f"translation_std_m>{max_tag_translation_std_m}")

        if max_tag_rotation_std_deg is not None:
            if rot_std > max_tag_rotation_std_deg:
                reasons.append(f"rotation_std_deg>{max_tag_rotation_std_deg}")

        if reasons:
            debug_info["removed"][tid] = {
                "reasons": reasons,
                "low_confidence": low_conf,
                "translation_std_m": trans_std,
                "rotation_std_deg": rot_std,
                "num_samples": int(item.get("num_samples", 0)),
            }
            continue

        tags[tid] = {
            "center_W": np.array(item["center_W"], dtype=np.float64),
            "num_samples": int(item.get("num_samples", 0)),
            "translation_std_m": trans_std,
            "rotation_std_deg": rot_std,
            "low_confidence": low_conf,
        }
        debug_info["kept_tag_ids"].append(tid)

    debug_info["all_tag_ids"] = sorted(debug_info["all_tag_ids"])
    debug_info["kept_tag_ids"] = sorted(debug_info["kept_tag_ids"])
    return tags, debug_info


def load_transforms_json(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return (
        np.array(data["T_W_C2"], dtype=np.float64),
        np.array(data["T_H_C1"], dtype=np.float64),
        np.array(data["T_H_HUCS"], dtype=np.float64),
    )


def transform_point(T_ab, p_b):
    return T_ab[:3, :3] @ p_b + T_ab[:3, 3]


def transform_dir(T_ab, d_b):
    v = T_ab[:3, :3] @ d_b
    n = np.linalg.norm(v)
    if n < 1e-12:
        return np.array([np.nan, np.nan, np.nan], dtype=np.float64)
    return v / n


def angle_between_deg(a, b):
    na = np.linalg.norm(a)
    nb = np.linalg.norm(b)
    if na < 1e-12 or nb < 1e-12:
        return np.nan
    a = a / na
    b = b / nb
    c = np.clip(np.dot(a, b), -1.0, 1.0)
    return float(np.degrees(np.arccos(c)))


def point_to_ray_distance(point, origin, direction):
    nd = np.linalg.norm(direction)
    if nd < 1e-12:
        return np.nan
    direction = direction / nd
    return float(np.linalg.norm(np.cross(point - origin, direction)))


def euler_zyx_deg_to_R(roll_deg, pitch_deg, yaw_deg):
    return R.from_euler("zyx", [yaw_deg, pitch_deg, roll_deg], degrees=True).as_matrix()


def build_T_C2_H(row):
    ok = to_int(row.get("ok"), default=0)
    if ok != 1:
        return None

    tx = to_float(row.get("cam_head_tx"))
    ty = to_float(row.get("cam_head_ty"))
    tz = to_float(row.get("cam_head_tz"))
    roll = to_float(row.get("cam_head_roll_deg"))
    pitch = to_float(row.get("cam_head_pitch_deg"))
    yaw = to_float(row.get("cam_head_yaw_deg"))

    if any(np.isnan(v) for v in [tx, ty, tz, roll, pitch, yaw]):
        return None

    Rm = euler_zyx_deg_to_R(roll, pitch, yaw)
    T = np.eye(4, dtype=np.float64)
    T[:3, :3] = Rm
    T[:3, 3] = [tx, ty, tz]
    return T


def nearest_row(rows, target_time_ns, time_col="unix_ns"):
    best = None
    best_dt = None
    for r in rows:
        t = to_int(r.get(time_col))
        if t is None:
            continue
        dt = abs(t - target_time_ns)
        if best_dt is None or dt < best_dt:
            best_dt = dt
            best = r
    return best, best_dt


def load_tobii_rows_raw(xlsx_path, recording_g3_path):
    recording_start_ns = get_recording_start_ns(recording_g3_path)

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) == 0:
        raise RuntimeError("Tobii xlsx is empty.")

    header = ["" if v is None else str(v).strip() for v in rows[0]]
    hidx = {k: i for i, k in enumerate(header)}

    processed = []

    for row in rows[1:]:
        comp_us = row[hidx["Computer timestamp"]] if "Computer timestamp" in hidx else None
        if comp_us is None:
            continue

        unix_ns = recording_start_ns + int(float(comp_us)) * 1000
        sensor = row[hidx["Sensor"]] if "Sensor" in hidx else ""

        if sensor != "Eye Tracker":
            continue

        rr = {"unix_ns": unix_ns}

        for k in [
            "Gaze point X", "Gaze point Y",
            "Gaze point 3D X", "Gaze point 3D Y", "Gaze point 3D Z",
            "Gaze direction left X", "Gaze direction left Y", "Gaze direction left Z",
            "Gaze direction right X", "Gaze direction right Y", "Gaze direction right Z",
            "Pupil position left X", "Pupil position left Y", "Pupil position left Z",
            "Pupil position right X", "Pupil position right Y", "Pupil position right Z",
            "Validity left", "Validity right",
        ]:
            rr[k] = row[hidx[k]] if k in hidx else None

        processed.append(rr)

    return processed


def summarize_vals(vals):
    vals = np.asarray(vals, dtype=np.float64)
    vals = vals[np.isfinite(vals)]
    if len(vals) == 0:
        return None
    return {
        "count": int(len(vals)),
        "mean": float(np.mean(vals)),
        "median": float(np.median(vals)),
        "p95": float(np.quantile(vals, 0.95)),
    }


def print_global_summary(name, out_rows, key):
    vals = []
    for r in out_rows:
        if key in r and r[key] != "" and r[key] is not None:
            try:
                vals.append(float(r[key]))
            except Exception:
                pass

    s = summarize_vals(vals)
    if s is None:
        print(f"{name}: no valid values")
        return

    print(
        f"{name}: n={s['count']}, "
        f"mean={s['mean']:.4f}, "
        f"median={s['median']:.4f}, "
        f"p95={s['p95']:.4f}"
    )


def dd_inc(d, key, amount=1):
    d[int(key)] += amount


def print_tag_counter(title, counter):
    print(title)
    if not counter:
        print("  none")
        return
    for k in sorted(counter.keys()):
        print(f"  tag {k}: {counter[k]}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tag-windows-csv", required=True)
    ap.add_argument("--apriltag-baseline-json", required=True)
    ap.add_argument("--headpose-csv", required=True)
    ap.add_argument("--scene-camera-json", required=True)
    ap.add_argument("--transforms-json", required=True)
    ap.add_argument("--tobii-raw-xlsx", required=True)
    ap.add_argument("--recording-g3", required=True)
    ap.add_argument("--output-csv", required=True)

    ap.add_argument("--window-fraction-start", type=float, default=0.3)
    ap.add_argument("--window-fraction-end", type=float, default=0.7)
    ap.add_argument("--max-sync-dt-ms", type=float, default=20.0)

    ap.add_argument("--b-require-both-eyes-valid", action="store_true")

    ap.add_argument("--exclude-low-confidence-tags", action="store_true")
    ap.add_argument("--max-tag-translation-std-m", type=float, default=None)
    ap.add_argument("--max-tag-rotation-std-deg", type=float, default=None)

    ap.add_argument("--min-head-tags", type=int, default=2)
    ap.add_argument("--max-head-rmse-px", type=float, default=5.0)

    ap.add_argument(
        "--debug-json",
        default=None,
        help="optional path to save debug counters as json",
    )

    args = ap.parse_args()

    windows = read_csv_dicts(args.tag_windows_csv)

    baseline, baseline_debug = load_baseline_json(
        args.apriltag_baseline_json,
        exclude_low_confidence=args.exclude_low_confidence_tags,
        max_tag_translation_std_m=args.max_tag_translation_std_m,
        max_tag_rotation_std_deg=args.max_tag_rotation_std_deg,
    )

    headpose_rows = read_csv_dicts(args.headpose_csv)
    K_scene, _ = load_camera_json(args.scene_camera_json)
    T_W_C2, T_H_C1, T_H_HUCS = load_transforms_json(args.transforms_json)
    tobii_rows = load_tobii_rows_raw(args.tobii_raw_xlsx, args.recording_g3)

    print("[INFO] input windows:", len(windows))
    print("[INFO] baseline all tags:", baseline_debug["all_tag_ids"])
    print("[INFO] baseline kept tags after filters:", baseline_debug["kept_tag_ids"])
    print("[INFO] headpose rows:", len(headpose_rows))
    print("[INFO] tobii eye-tracker rows:", len(tobii_rows))

    debug = {
        "input_windows_by_tag": defaultdict(int),
        "skipped_invalid_window": defaultdict(int),
        "skipped_tag_not_in_baseline": defaultdict(int),
        "windows_with_tobii_rows": defaultdict(int),
        "windows_without_tobii_rows": defaultdict(int),
        "tobii_rows_in_windows_by_tag": defaultdict(int),
        "rows_skipped_no_headpose_match": defaultdict(int),
        "rows_skipped_headpose_quality": defaultdict(int),
        "rows_skipped_invalid_headpose_transform": defaultdict(int),
        "out_rows_by_tag": defaultdict(int),
        "A_rows_by_tag": defaultdict(int),
        "B_total_checked_by_tag": defaultdict(int),
        "B_valid_left_by_tag": defaultdict(int),
        "B_valid_right_by_tag": defaultdict(int),
        "B_both_valid_by_tag": defaultdict(int),
        "B_used_rows_by_tag": defaultdict(int),
        "B_metric_rows_by_tag": defaultdict(int),
        "C_rows_by_tag": defaultdict(int),
    }

    out_rows = []

    count_B_total_checked = 0
    count_B_valid_left = 0
    count_B_valid_right = 0
    count_B_both_valid = 0
    count_B_used_rows = 0

    for w in windows:
        tag_id = to_int(w.get("tag_id"))
        start_ts = to_int(w.get("start_unix_ns"))
        end_ts = to_int(w.get("end_unix_ns"))
        window_id = to_int(w.get("window_id"))

        if None in [tag_id, start_ts, end_ts, window_id]:
            if tag_id is not None:
                dd_inc(debug["skipped_invalid_window"], tag_id)
            continue

        dd_inc(debug["input_windows_by_tag"], tag_id)

        if tag_id not in baseline:
            dd_inc(debug["skipped_tag_not_in_baseline"], tag_id)
            continue

        target_W = baseline[tag_id]["center_W"]

        dur = end_ts - start_ts
        if dur <= 0:
            dd_inc(debug["skipped_invalid_window"], tag_id)
            continue

        sub_start_ts = int(start_ts + dur * args.window_fraction_start)
        sub_end_ts = int(start_ts + dur * args.window_fraction_end)

        num_tobii_in_window = 0
        num_out_rows_this_window = 0

        for tr in tobii_rows:
            t_ns = tr["unix_ns"]
            if not (sub_start_ts <= t_ns <= sub_end_ts):
                continue

            num_tobii_in_window += 1
            dd_inc(debug["tobii_rows_in_windows_by_tag"], tag_id)

            head_row, dt_ns = nearest_row(headpose_rows, t_ns, time_col="timestamp_ns")
            if head_row is None or dt_ns is None or dt_ns / 1e6 > args.max_sync_dt_ms:
                dd_inc(debug["rows_skipped_no_headpose_match"], tag_id)
                continue

            num_head_tags = to_int(head_row.get("num_head_tags"), default=0)
            head_rmse_px = to_float(head_row.get("head_rmse_px"))

            if num_head_tags < args.min_head_tags:
                dd_inc(debug["rows_skipped_headpose_quality"], tag_id)
                continue

            if np.isnan(head_rmse_px) or head_rmse_px > args.max_head_rmse_px:
                dd_inc(debug["rows_skipped_headpose_quality"], tag_id)
                continue

            T_C2_H = build_T_C2_H(head_row)
            if T_C2_H is None:
                dd_inc(debug["rows_skipped_invalid_headpose_transform"], tag_id)
                continue

            T_W_H = T_W_C2 @ T_C2_H

            row_out = {
                "window_id": window_id,
                "tag_id": tag_id,
                "tobii_unix_ns": t_ns,
                "headpose_dt_ms": dt_ns / 1e6,
                "head_num_tags": num_head_tags,
                "head_rmse_px": head_rmse_px,
            }

            # A: Tobii native 3D gaze point
            gx3 = to_float(tr.get("Gaze point 3D X"))
            gy3 = to_float(tr.get("Gaze point 3D Y"))
            gz3 = to_float(tr.get("Gaze point 3D Z"))

            if not any(np.isnan(v) for v in [gx3, gy3, gz3]):
                p_hucs = np.array([gx3, gy3, gz3], dtype=np.float64) * 1e-3
                p_H = transform_point(T_H_HUCS, p_hucs)
                p_W = transform_point(T_W_H, p_H)
                row_out["A_position_error_m"] = float(np.linalg.norm(p_W - target_W))
                dd_inc(debug["A_rows_by_tag"], tag_id)

            # B: Tobii native 3D gaze ray
            count_B_total_checked += 1
            dd_inc(debug["B_total_checked_by_tag"], tag_id)

            vl = str(tr.get("Validity left") or "").strip()
            vr = str(tr.get("Validity right") or "").strip()

            pl = np.array([
                to_float(tr.get("Pupil position left X")),
                to_float(tr.get("Pupil position left Y")),
                to_float(tr.get("Pupil position left Z")),
            ], dtype=np.float64) * 1e-3

            pr = np.array([
                to_float(tr.get("Pupil position right X")),
                to_float(tr.get("Pupil position right Y")),
                to_float(tr.get("Pupil position right Z")),
            ], dtype=np.float64) * 1e-3

            dl = np.array([
                to_float(tr.get("Gaze direction left X")),
                to_float(tr.get("Gaze direction left Y")),
                to_float(tr.get("Gaze direction left Z")),
            ], dtype=np.float64)

            dr = np.array([
                to_float(tr.get("Gaze direction right X")),
                to_float(tr.get("Gaze direction right Y")),
                to_float(tr.get("Gaze direction right Z")),
            ], dtype=np.float64)

            valid_l = np.all(np.isfinite(pl)) and np.all(np.isfinite(dl)) and vl == "Valid"
            valid_r = np.all(np.isfinite(pr)) and np.all(np.isfinite(dr)) and vr == "Valid"

            if valid_l:
                count_B_valid_left += 1
                dd_inc(debug["B_valid_left_by_tag"], tag_id)

            if valid_r:
                count_B_valid_right += 1
                dd_inc(debug["B_valid_right_by_tag"], tag_id)

            if valid_l and valid_r:
                count_B_both_valid += 1
                dd_inc(debug["B_both_valid_by_tag"], tag_id)

            if args.b_require_both_eyes_valid:
                use_B = valid_l and valid_r
            else:
                use_B = valid_l or valid_r

            if use_B:
                count_B_used_rows += 1
                dd_inc(debug["B_used_rows_by_tag"], tag_id)

                if valid_l and valid_r:
                    o_hucs = 0.5 * (pl + pr)
                    d_hucs = 0.5 * (dl + dr)
                elif valid_l:
                    o_hucs = pl
                    d_hucs = dl
                else:
                    o_hucs = pr
                    d_hucs = dr

                nd = np.linalg.norm(d_hucs)
                if nd > 1e-12:
                    d_hucs = d_hucs / nd
                    o_H = transform_point(T_H_HUCS, o_hucs)
                    d_H = transform_dir(T_H_HUCS, d_hucs)
                    o_W = transform_point(T_W_H, o_H)
                    d_W = transform_dir(T_W_H, d_H)

                    d_gt = target_W - o_W
                    ndgt = np.linalg.norm(d_gt)

                    if ndgt > 1e-12 and np.all(np.isfinite(d_W)):
                        d_gt = d_gt / ndgt
                        row_out["B_angle_error_deg"] = float(angle_between_deg(d_W, d_gt))
                        row_out["B_point_to_ray_m"] = float(
                            point_to_ray_distance(target_W, o_W, d_W)
                        )
                        dd_inc(debug["B_metric_rows_by_tag"], tag_id)

            # C: 2D gaze point + scene camera intrinsics
            gx = to_float(tr.get("Gaze point X"))
            gy = to_float(tr.get("Gaze point Y"))

            if not any(np.isnan(v) for v in [gx, gy]):
                uv1 = np.array([gx, gy, 1.0], dtype=np.float64)
                d_C1 = np.linalg.inv(K_scene) @ uv1
                nd = np.linalg.norm(d_C1)

                if nd > 1e-12:
                    d_C1 = d_C1 / nd
                    o_C1 = np.zeros(3, dtype=np.float64)

                    o_H = transform_point(T_H_C1, o_C1)
                    d_H = transform_dir(T_H_C1, d_C1)
                    o_W = transform_point(T_W_H, o_H)
                    d_W = transform_dir(T_W_H, d_H)

                    d_gt = target_W - o_W
                    ndgt = np.linalg.norm(d_gt)

                    if ndgt > 1e-12 and np.all(np.isfinite(d_W)):
                        d_gt = d_gt / ndgt
                        row_out["C_angle_error_deg"] = float(angle_between_deg(d_W, d_gt))
                        row_out["C_point_to_ray_m"] = float(
                            point_to_ray_distance(target_W, o_W, d_W)
                        )
                        dd_inc(debug["C_rows_by_tag"], tag_id)

            out_rows.append(row_out)
            num_out_rows_this_window += 1
            dd_inc(debug["out_rows_by_tag"], tag_id)

        if num_tobii_in_window > 0:
            dd_inc(debug["windows_with_tobii_rows"], tag_id)
        else:
            dd_inc(debug["windows_without_tobii_rows"], tag_id)

    # ------------------------------------------------------------
    # Debug print before failure
    # ------------------------------------------------------------
    print("\n=== DEBUG: windows / filtering by tag ===")
    print_tag_counter("input_windows_by_tag", debug["input_windows_by_tag"])
    print_tag_counter("skipped_tag_not_in_baseline", debug["skipped_tag_not_in_baseline"])
    print_tag_counter("windows_with_tobii_rows", debug["windows_with_tobii_rows"])
    print_tag_counter("windows_without_tobii_rows", debug["windows_without_tobii_rows"])
    print_tag_counter("tobii_rows_in_windows_by_tag", debug["tobii_rows_in_windows_by_tag"])
    print_tag_counter("rows_skipped_no_headpose_match", debug["rows_skipped_no_headpose_match"])
    print_tag_counter("rows_skipped_headpose_quality", debug["rows_skipped_headpose_quality"])
    print_tag_counter("rows_skipped_invalid_headpose_transform", debug["rows_skipped_invalid_headpose_transform"])
    print_tag_counter("out_rows_by_tag", debug["out_rows_by_tag"])
    print_tag_counter("A_rows_by_tag", debug["A_rows_by_tag"])
    print_tag_counter("B_metric_rows_by_tag", debug["B_metric_rows_by_tag"])
    print_tag_counter("C_rows_by_tag", debug["C_rows_by_tag"])

    if args.debug_json:
        dbg_out = {
            "baseline_debug": {
                "all_tag_ids": baseline_debug["all_tag_ids"],
                "kept_tag_ids": baseline_debug["kept_tag_ids"],
                "removed": {str(k): v for k, v in baseline_debug["removed"].items()},
            },
            "counters": {
                name: {str(k): int(v) for k, v in counter.items()}
                for name, counter in debug.items()
            },
            "settings": vars(args),
        }

        out_dbg = Path(args.debug_json)
        out_dbg.parent.mkdir(parents=True, exist_ok=True)
        with open(out_dbg, "w", encoding="utf-8") as f:
            json.dump(dbg_out, f, indent=2)
        print(f"[INFO] saved debug json to {out_dbg}")

    if not out_rows:
        raise RuntimeError("No valid A/B/C rows produced.")

    # ------------------------------------------------------------
    # Save per-row CSV
    # ------------------------------------------------------------
    fieldnames = []
    seen = set()

    for r in out_rows:
        for k in r:
            if k not in seen:
                seen.add(k)
                fieldnames.append(k)

    output_path = Path(args.output_csv)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(out_rows)

    print(f"saved to {args.output_csv}")

    # ------------------------------------------------------------
    # Summary CSV
    # ------------------------------------------------------------
    grouped = defaultdict(list)
    for r in out_rows:
        key = (int(r["window_id"]), int(r["tag_id"]))
        grouped[key].append(r)

    summary_rows = []

    for (window_id, tag_id), rows_g in grouped.items():
        row_sum = {
            "window_id": window_id,
            "tag_id": tag_id,
            "num_rows": len(rows_g),
        }

        s = summarize_vals([
            float(r["headpose_dt_ms"])
            for r in rows_g
            if "headpose_dt_ms" in r and r["headpose_dt_ms"] != ""
        ])
        if s is not None:
            row_sum["headpose_dt_ms_mean"] = s["mean"]
            row_sum["headpose_dt_ms_median"] = s["median"]
            row_sum["headpose_dt_ms_p95"] = s["p95"]

        s = summarize_vals([
            float(r["A_position_error_m"])
            for r in rows_g
            if "A_position_error_m" in r and r["A_position_error_m"] != ""
        ])
        if s is not None:
            row_sum["A_position_error_m_mean"] = s["mean"]
            row_sum["A_position_error_m_median"] = s["median"]
            row_sum["A_position_error_m_p95"] = s["p95"]

        s = summarize_vals([
            float(r["B_angle_error_deg"])
            for r in rows_g
            if "B_angle_error_deg" in r and r["B_angle_error_deg"] != ""
        ])
        if s is not None:
            row_sum["B_angle_error_deg_mean"] = s["mean"]
            row_sum["B_angle_error_deg_median"] = s["median"]
            row_sum["B_angle_error_deg_p95"] = s["p95"]

        s = summarize_vals([
            float(r["B_point_to_ray_m"])
            for r in rows_g
            if "B_point_to_ray_m" in r and r["B_point_to_ray_m"] != ""
        ])
        if s is not None:
            row_sum["B_point_to_ray_m_mean"] = s["mean"]
            row_sum["B_point_to_ray_m_median"] = s["median"]
            row_sum["B_point_to_ray_m_p95"] = s["p95"]

        s = summarize_vals([
            float(r["C_angle_error_deg"])
            for r in rows_g
            if "C_angle_error_deg" in r and r["C_angle_error_deg"] != ""
        ])
        if s is not None:
            row_sum["C_angle_error_deg_mean"] = s["mean"]
            row_sum["C_angle_error_deg_median"] = s["median"]
            row_sum["C_angle_error_deg_p95"] = s["p95"]

        s = summarize_vals([
            float(r["C_point_to_ray_m"])
            for r in rows_g
            if "C_point_to_ray_m" in r and r["C_point_to_ray_m"] != ""
        ])
        if s is not None:
            row_sum["C_point_to_ray_m_mean"] = s["mean"]
            row_sum["C_point_to_ray_m_median"] = s["median"]
            row_sum["C_point_to_ray_m_p95"] = s["p95"]

        summary_rows.append(row_sum)

    summary_csv = str(output_path.with_name(output_path.stem + "_summary.csv"))

    if len(summary_rows) > 0:
        sf = []
        seen = set()
        for r in summary_rows:
            for k in r.keys():
                if k not in seen:
                    seen.add(k)
                    sf.append(k)

        with open(summary_csv, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=sf)
            writer.writeheader()
            writer.writerows(summary_rows)

        print(f"saved summary to {summary_csv}")

    # ------------------------------------------------------------
    # Global summary
    # ------------------------------------------------------------
    print_global_summary("A_position_error_m", out_rows, "A_position_error_m")
    print_global_summary("B_angle_error_deg", out_rows, "B_angle_error_deg")
    print_global_summary("B_point_to_ray_m", out_rows, "B_point_to_ray_m")
    print_global_summary("C_angle_error_deg", out_rows, "C_angle_error_deg")
    print_global_summary("C_point_to_ray_m", out_rows, "C_point_to_ray_m")
    print_global_summary("headpose_dt_ms", out_rows, "headpose_dt_ms")

    print(f"B total checked = {count_B_total_checked}")
    print(f"B valid left = {count_B_valid_left}")
    print(f"B valid right = {count_B_valid_right}")
    print(f"B both valid = {count_B_both_valid}")
    print(f"B used rows = {count_B_used_rows}")

    print("\n=== DEBUG: B validity by tag ===")
    print_tag_counter("B_total_checked_by_tag", debug["B_total_checked_by_tag"])
    print_tag_counter("B_valid_left_by_tag", debug["B_valid_left_by_tag"])
    print_tag_counter("B_valid_right_by_tag", debug["B_valid_right_by_tag"])
    print_tag_counter("B_both_valid_by_tag", debug["B_both_valid_by_tag"])
    print_tag_counter("B_used_rows_by_tag", debug["B_used_rows_by_tag"])

    gc.collect()
    sys.stdout.flush()
    os._exit(0)


if __name__ == "__main__":
    main()