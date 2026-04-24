#!/usr/bin/env python3
import argparse
import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from collections import defaultdict
import os
import sys
import gc

import cv2
import numpy as np
from openpyxl import load_workbook
from pupil_apriltags import Detector


def to_int(v, default=None):
    try:
        if v is None or v == "":
            return default
        return int(float(v))
    except Exception:
        return default


def to_float(v, default=np.nan):
    try:
        if v is None or v == "":
            return default
        return float(v)
    except Exception:
        return default


def read_csv_dicts(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def iso_to_unix_ns(iso_str: str) -> int:
    s = iso_str.strip()
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    dt = datetime.fromisoformat(s)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return int(round(dt.timestamp() * 1e9))


def get_recording_start_ns(recording_g3_path):
    with open(recording_g3_path, "r", encoding="utf-8") as f:
        g3 = json.load(f)
    return iso_to_unix_ns(g3["created"])


def point_in_polygon(pt, polygon):
    polygon = np.asarray(polygon, dtype=np.float32)
    pt = (float(pt[0]), float(pt[1]))
    return cv2.pointPolygonTest(polygon, pt, False) >= 0


def point_to_polygon_distance(pt, polygon):
    polygon = np.asarray(polygon, dtype=np.float32)
    pt = (float(pt[0]), float(pt[1]))
    signed = cv2.pointPolygonTest(polygon, pt, True)
    if signed >= 0:
        return 0.0
    return abs(float(signed))


def reorder_corners(corners):
    c = np.asarray(corners, dtype=np.float32).reshape(4, 2)
    return c[[3, 2, 1, 0], :]


def nearest_row(rows, target_time_ns, time_col="unix_ns"):
    best = None
    best_dt = None
    for r in rows:
        t = to_int(r.get(time_col))
        if t is None:
            continue
        dt = abs(t - target_time_ns)
        if best_dt is None or dt < best_dt:
            best_dt = dt
            best = r
    return best, best_dt


def resolve_frame_path(frame_dir, frame_idx):
    frame_idx = int(frame_idx)
    for ext in [".png", ".jpg", ".jpeg"]:
        p = frame_dir / f"frame_{frame_idx:06d}{ext}"
        if p.exists():
            return p
    return None


def preprocess_tobii_rows_xlsx(tobii_xlsx_path, recording_g3_path, mode):
    recording_start_ns = get_recording_start_ns(recording_g3_path)

    wb = load_workbook(tobii_xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = ["" if v is None else str(v).strip() for v in rows[0]]
    hidx = {k: i for i, k in enumerate(header)}

    processed = []
    for row in rows[1:]:
        comp_us = row[hidx["Computer timestamp"]] if "Computer timestamp" in hidx else None
        if comp_us is None:
            continue
        unix_ns = recording_start_ns + int(float(comp_us)) * 1000

        if mode == "raw":
            sensor = row[hidx["Sensor"]] if "Sensor" in hidx else ""
            gx = to_float(row[hidx["Gaze point X"]]) if "Gaze point X" in hidx else np.nan
            gy = to_float(row[hidx["Gaze point Y"]]) if "Gaze point Y" in hidx else np.nan
            if sensor != "Eye Tracker":
                continue
            if np.isnan(gx) or np.isnan(gy):
                continue
            processed.append({"unix_ns": unix_ns, "_gaze_x": gx, "_gaze_y": gy, "_gaze_source": "raw"})

        elif mode == "fixation":
            emt = row[hidx["Eye movement type"]] if "Eye movement type" in hidx else ""
            fx = to_float(row[hidx["Fixation point X"]]) if "Fixation point X" in hidx else np.nan
            fy = to_float(row[hidx["Fixation point Y"]]) if "Fixation point Y" in hidx else np.nan
            if emt != "Fixation":
                continue
            if np.isnan(fx) or np.isnan(fy):
                continue
            processed.append({"unix_ns": unix_ns, "_gaze_x": fx, "_gaze_y": fy, "_gaze_source": "fixation"})

        else:
            raise ValueError("mode must be raw or fixation")

    return processed, recording_start_ns


def summarize_vals(vals):
    vals = np.asarray(vals, dtype=np.float64)
    vals = vals[np.isfinite(vals)]
    if len(vals) == 0:
        return None
    return {
        "count": int(len(vals)),
        "mean": float(np.mean(vals)),
        "median": float(np.median(vals)),
        "p95": float(np.quantile(vals, 0.95)),
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tag-windows-csv", required=True)
    ap.add_argument("--scene-timestamps-csv", required=True)
    ap.add_argument("--scene-frame-dir", required=True)
    ap.add_argument("--tobii-xlsx", required=True)
    ap.add_argument("--recording-g3", required=True)
    ap.add_argument("--mode", required=True, choices=["raw", "fixation"])
    ap.add_argument("--output-csv", required=True)
    ap.add_argument("--tag-family", default="tag36h11")
    ap.add_argument("--max-tobii-dt-ms", type=float, default=100.0)
    ap.add_argument("--window-fraction-start", type=float, default=0.0)
    ap.add_argument("--window-fraction-end", type=float, default=1.0)
    args = ap.parse_args()

    windows = read_csv_dicts(args.tag_windows_csv)
    scene_rows = read_csv_dicts(args.scene_timestamps_csv)
    tobii_rows, recording_start_ns = preprocess_tobii_rows_xlsx(
        args.tobii_xlsx, args.recording_g3, args.mode
    )

    print(f"[INFO] recording_start_ns = {recording_start_ns}")
    print(f"[INFO] processed Tobii rows ({args.mode}) = {len(tobii_rows)}")

    scene_by_frame = {}
    for r in scene_rows:
        fi = to_int(r.get("frame_idx"))
        if fi is not None:
            scene_by_frame[fi] = r

    detector = Detector(
        families=args.tag_family,
        nthreads=4,
        quad_decimate=1.0,
        quad_sigma=0.0,
        refine_edges=1,
        decode_sharpening=0.25,
        debug=0,
    )

    frame_dir = Path(args.scene_frame_dir)
    out_rows = []

    for w in windows:
        window_id = to_int(w.get("window_id"))
        tag_id = to_int(w.get("tag_id"))
        start_f = to_int(w.get("start_frame_idx"))
        end_f = to_int(w.get("end_frame_idx"))

        if None in [window_id, tag_id, start_f, end_f]:
            continue

        n_frames = end_f - start_f + 1
        sub_start = start_f + int(np.floor(n_frames * args.window_fraction_start))
        sub_end = start_f + int(np.ceil(n_frames * args.window_fraction_end)) - 1
        sub_start = max(sub_start, start_f)
        sub_end = min(sub_end, end_f)
        if sub_end < sub_start:
            continue

        for frame_idx in range(sub_start, sub_end + 1):
            if frame_idx not in scene_by_frame:
                continue

            sr = scene_by_frame[frame_idx]
            scene_unix_ns = to_int(sr.get("unix_ns"))
            if scene_unix_ns is None:
                continue

            img_path = resolve_frame_path(frame_dir, frame_idx)
            if img_path is None:
                out_rows.append({
                    "window_id": window_id,
                    "tag_id": tag_id,
                    "frame_idx": frame_idx,
                    "status": "scene_image_missing"
                })
                continue

            img = cv2.imread(str(img_path), cv2.IMREAD_COLOR)
            if img is None:
                out_rows.append({
                    "window_id": window_id,
                    "tag_id": tag_id,
                    "frame_idx": frame_idx,
                    "status": "scene_image_read_failed"
                })
                continue

            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            dets = detector.detect(gray, estimate_tag_pose=False)

            target_det = None
            for d in dets:
                if int(d.tag_id) == tag_id:
                    target_det = d
                    break

            if target_det is None:
                out_rows.append({
                    "window_id": window_id,
                    "tag_id": tag_id,
                    "frame_idx": frame_idx,
                    "status": "target_tag_not_detected"
                })
                continue

            polygon = reorder_corners(target_det.corners)
            center = np.mean(polygon, axis=0)

            widths = [
                np.linalg.norm(polygon[1] - polygon[0]),
                np.linalg.norm(polygon[2] - polygon[3]),
            ]
            heights = [
                np.linalg.norm(polygon[2] - polygon[1]),
                np.linalg.norm(polygon[3] - polygon[0]),
            ]
            tag_width_px = float(np.mean(widths))
            tag_height_px = float(np.mean(heights))

            tobii_row, dt_ns = nearest_row(tobii_rows, scene_unix_ns, time_col="unix_ns")
            if tobii_row is None:
                out_rows.append({
                    "window_id": window_id,
                    "tag_id": tag_id,
                    "frame_idx": frame_idx,
                    "status": "tobii_missing"
                })
                continue

            dt_ms = dt_ns / 1e6
            if dt_ms > args.max_tobii_dt_ms:
                out_rows.append({
                    "window_id": window_id,
                    "tag_id": tag_id,
                    "frame_idx": frame_idx,
                    "status": "tobii_too_far",
                    "nearest_tobii_dt_ms": dt_ms
                })
                continue

            gx = float(tobii_row["_gaze_x"])
            gy = float(tobii_row["_gaze_y"])
            gaze_source = tobii_row["_gaze_source"]
            gaze_pt = np.array([gx, gy], dtype=np.float32)

            inside = point_in_polygon(gaze_pt, polygon)
            center_err_px = float(np.linalg.norm(gaze_pt - center))
            poly_dist_px = float(point_to_polygon_distance(gaze_pt, polygon))

            out_rows.append({
                "window_id": window_id,
                "tag_id": tag_id,
                "frame_idx": frame_idx,
                "scene_unix_ns": scene_unix_ns,
                "tobii_unix_ns": int(tobii_row["unix_ns"]),
                "nearest_tobii_dt_ms": dt_ms,
                "gaze_source": gaze_source,
                "status": "ok",
                "gaze_x": gx,
                "gaze_y": gy,
                "tag_center_x": float(center[0]),
                "tag_center_y": float(center[1]),
                "inside_tag_polygon": int(inside),
                "center_error_px": center_err_px,
                "distance_to_polygon_px": poly_dist_px,
                "tag_width_px": tag_width_px,
                "tag_height_px": tag_height_px,
                "center_error_over_tag_width": float(center_err_px / tag_width_px) if tag_width_px > 1e-9 else np.nan,
                "distance_to_polygon_over_tag_width": float(poly_dist_px / tag_width_px) if tag_width_px > 1e-9 else np.nan,
            })

    if len(out_rows) == 0:
        print("No output rows.")
        return

    fieldnames = []
    seen = set()
    for r in out_rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                fieldnames.append(k)

    with open(args.output_csv, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(out_rows)

    print(f"[INFO] saved to {args.output_csv}")

    # summary csv by window/tag
    ok_rows = [r for r in out_rows if r.get("status") == "ok"]
    grouped = defaultdict(list)
    for r in ok_rows:
        grouped[(int(r["window_id"]), int(r["tag_id"]))].append(r)

    summary_rows = []
    for (window_id, tag_id), rows_g in grouped.items():
        row = {
            "window_id": window_id,
            "tag_id": tag_id,
            "num_valid_rows": len(rows_g),
        }

        inside_vals = np.array([int(r["inside_tag_polygon"]) for r in rows_g], dtype=np.int32)
        row["inside_rate"] = float(np.mean(inside_vals))

        for key in [
            "center_error_px",
            "distance_to_polygon_px",
            "tag_width_px",
            "tag_height_px",
            "center_error_over_tag_width",
            "distance_to_polygon_over_tag_width",
            "nearest_tobii_dt_ms",
        ]:
            s = summarize_vals([float(r[key]) for r in rows_g if key in r and r[key] != ""])
            if s is not None:
                row[f"{key}_mean"] = s["mean"]
                row[f"{key}_median"] = s["median"]
                row[f"{key}_p95"] = s["p95"]

        summary_rows.append(row)

    summary_csv = str(Path(args.output_csv).with_name(Path(args.output_csv).stem + "_summary.csv"))
    if summary_rows:
        sf = []
        seen = set()
        for r in summary_rows:
            for k in r.keys():
                if k not in seen:
                    seen.add(k)
                    sf.append(k)

        with open(summary_csv, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=sf)
            writer.writeheader()
            writer.writerows(summary_rows)

        print(f"[INFO] saved summary to {summary_csv}")

    # global summary
    if len(ok_rows) > 0:
        inside_vals = np.array([int(r["inside_tag_polygon"]) for r in ok_rows], dtype=np.int32)
        center_err = np.array([float(r["center_error_px"]) for r in ok_rows], dtype=np.float64)
        poly_dist = np.array([float(r["distance_to_polygon_px"]) for r in ok_rows], dtype=np.float64)
        dt_vals = np.array([float(r["nearest_tobii_dt_ms"]) for r in ok_rows], dtype=np.float64)
        ratio = np.array([float(r["distance_to_polygon_over_tag_width"]) for r in ok_rows], dtype=np.float64)

        print(f"[INFO] valid_rows = {len(ok_rows)}")
        print(f"[INFO] inside_rate = {inside_vals.mean():.4f}")
        print(f"[INFO] center_error_px: mean={center_err.mean():.2f}, median={np.median(center_err):.2f}, p95={np.quantile(center_err,0.95):.2f}")
        print(f"[INFO] distance_to_polygon_px: mean={poly_dist.mean():.2f}, median={np.median(poly_dist):.2f}, p95={np.quantile(poly_dist,0.95):.2f}")
        print(f"[INFO] distance_to_polygon_over_tag_width: mean={ratio.mean():.2f}, median={np.median(ratio):.2f}, p95={np.quantile(ratio,0.95):.2f}")
        print(f"[INFO] nearest_tobii_dt_ms: mean={dt_vals.mean():.2f}, median={np.median(dt_vals):.2f}, p95={np.quantile(dt_vals,0.95):.2f}")

    gc.collect()
    sys.stdout.flush()
    os._exit(0)


if __name__ == "__main__":
    main()